# 合并、拆分单元格
# 合并单元格有下面两种方法，需要注意的是，如果要合并的格子中有数据，即便python没有报错，Excel打开的时候也会报错。
# merge_cells(待合并的格子编号)
# merge_cells(start_row=起始行号，start_column=起始列号，end_row=结束行号，end_column=结束列号)
import os
import openpyxl
import openpyxl.styles

path = r"C:\Users\AMD\Desktop"
os.chdir(path)  

workbook = openpyxl.load_workbook('test.xlsx') 
sheet = workbook.active  
print('当前活动表是：' + str(sheet))

# 方法1：
sheet.merge_cells('A12:B13')
# 方法2：
sheet.merge_cells(start_row=12, start_column=3, end_row=13, end_column=4)

# 加一个居中对齐
cell = sheet['A12']
alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
cell.alignment = alignment

cell = sheet['C12']
alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
cell.alignment = alignment

workbook.save('test.xlsx')

# 拆分单元格的方法同上
# unmerge_cells(待拆分的格子编号)
# unmerge_cells(start_row=起始行号，start_column=起始列号，end_row=结束行号，end_column=结束列号)