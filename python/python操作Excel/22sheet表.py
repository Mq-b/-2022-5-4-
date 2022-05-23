# create_sheet(“新的sheet名”)：创建一个新的sheet表
import os
import openpyxl

path = r"C:\Users\AMD\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')  # 返回一个workbook数据类型的值
sheet = workbook.active  # 获取活动表
print('当前活动表是：' + str(sheet))

workbook.create_sheet('3号sheet')    # 创建新的新表
print(workbook.sheetnames)			 # 查看所有的表

# 修改表名字(title)
sheet.title='修改表名字sheet'
# 复制表
workbook.copy_worksheet(sheet)
#修改拷贝表的名字
sheet=workbook['Sheet1 Copy']
sheet.title='复制表'

#删除表
workbook.remove(workbook['修改表名字sheet'])

workbook.save('test.xlsx')