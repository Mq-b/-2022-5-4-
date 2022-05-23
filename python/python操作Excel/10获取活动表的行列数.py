import os
import openpyxl

path=r"C:\Users\AMD\Desktop"
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test.xlsx')  #
sheet = workbook.active  # 获取活动表
print('当前活动表是：' + str(sheet))

rows = sheet.max_row        # 获取行数
column = sheet.max_column   # 获取列数

print(rows)
print(column)
# 在处理Excel表格有时可能需要对表格进行遍历查找，openpyxl中便提供了一个行和列的
# 生成器(sheet.rows和sheet.columns),这两个生成器里面是每一行（或列）的数据
# ，每一行（或列）又由一个tuple包裹，借此可以很方便地完成对行和列的遍历