# row_dimensions[行编号].height = 行高
# column_dimensions[列编号].width = 列宽
import os
import openpyxl
import openpyxl.styles
from openpyxl.utils import get_column_letter
path=r"C:\Users\AMD\Desktop"
os.chdir(path) 

workbook = openpyxl.load_workbook('test.xlsx') 
sheet = workbook.active 
print('当前活动表是：' + str(sheet))

# 设置第1行的高度
sheet.row_dimensions[1].height = 50
# 设置B列的宽度
sheet.column_dimensions['B'].width = 20

# 设置所有单元格
for i in range(1,sheet.max_row+1):
    sheet.row_dimensions[i].height=20
for j in range(1,sheet.max_column+1):
    sheet.column_dimensions[get_column_letter(j)].width=30
workbook.save('test.xlsx')
