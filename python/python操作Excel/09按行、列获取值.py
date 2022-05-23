# iter_rows():按行读取
# iter_cols():按列读取
import os
import openpyxl

path=r"C:\Users\AMD\Desktop"
os.chdir(path) 

workbook=openpyxl.load_workbook('test.xlsx')
sheet =workbook.active
print('当前活动表是:'+str(sheet))

print('按行获取值')
for i in sheet.iter_rows(min_row=2,max_row=5,min_col=1,max_col=2):
    for j in i:
        print(j.value)

print('按列获取值')
for i in sheet.iter_cols(min_row=2,max_row=5,min_col=1,max_col=2):
    for j in i:
        print(j.value)
# min_row表示开始行数max_row表示结束行数